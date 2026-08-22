from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

wb = Workbook()
ws = wb.active
ws.title = "Performance"

widths = {
    "A": 10, "B": 28, "C": 28, "D": 28, "E": 28,
    "F": 16, "G": 16, "H": 16, "I": 14, "J": 14,
    "K": 22, "L": 22,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
thick = Border(
    left=Side(style="medium", color="000000"),
    right=Side(style="medium", color="000000"),
    top=Side(style="medium", color="000000"),
    bottom=Side(style="medium", color="000000"),
)

fill_header = PatternFill("solid", fgColor="D9E2F3")
fill_title = PatternFill("solid", fgColor="FFF2CC")
fill_scenario = PatternFill("solid", fgColor="FCE4D6")
fill_prereq_header = PatternFill("solid", fgColor="E2EFDA")
fill_green = PatternFill("solid", fgColor="92D050")
fill_orange = PatternFill("solid", fgColor="F4B183")
fill_red = PatternFill("solid", fgColor="FF5050")
fill_gray = PatternFill("solid", fgColor="F2F2F2")

font_bold = Font(bold=True, name="Calibri", size=11)
font_title = Font(bold=True, name="Calibri", size=14)
font_normal = Font(name="Calibri", size=11)
wrap = Alignment(wrap_text=True, vertical="center")
center = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_range(cell_range, fill=None, border=thin, alignment=wrap):
    from openpyxl.utils import range_boundaries

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = alignment
            if fill:
                cell.fill = fill


# Header
ws.merge_cells("A1:C1")
ws["A1"] = "Test Case Description"
ws["A1"].font = font_bold
ws["A1"].fill = fill_header
ws["A1"].alignment = center
style_range("A1:C1", fill_header, thin, center)

ws.merge_cells("D1:L2")
ws["D1"] = "Test the functionality of Performance Management Module"
ws["D1"].font = Font(bold=True, name="Calibri", size=16)
ws["D1"].fill = fill_title
ws["D1"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
style_range("D1:L2", fill_title, thin, Alignment(wrap_text=True, vertical="center", horizontal="left"))

ws.merge_cells("A4:C4")
ws["A4"] = "UAT Tester's Log"
ws["A4"].font = font_title

ws["A6"] = "Tester's Name"
ws["A6"].font = font_bold
ws["A6"].fill = fill_gray
ws["A6"].border = thin
ws.merge_cells("B6:C6")
ws["B6"].border = thick
style_range("B6:C6", border=thick)

ws["E6"] = "Date Tested"
ws["E6"].font = font_bold
ws["E6"].fill = fill_gray
ws["E6"].border = thin
ws.merge_cells("F6:G6")
style_range("F6:G6", border=thick)

# Prerequisites
ws["A8"] = "S #"
ws["B8"] = "Prerequisites:"
ws["A8"].font = font_bold
ws["B8"].font = font_bold
ws.merge_cells("B8:C8")
style_range("A8:C8", fill_prereq_header, thin, center)
ws["B8"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")

prereqs = [
    "Access to an internet browser",
    "Access to the internet",
    "Registered Email address / active StawiHR user accounts (Admin/HR, Supervisor, Staff)",
    "Login link: https://demo.stawihr.com/  (or local: http://localhost:8081/)",
    "Performance Management module enabled for the organization",
    "At least one active Employee with a Supervisor assigned",
    "At least one Review Period configured (or ability to create one)",
]
for i, text in enumerate(prereqs, start=1):
    row = 8 + i
    ws[f"A{row}"] = i
    ws[f"A{row}"].alignment = center
    ws[f"A{row}"].border = thin
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = text
    ws[f"B{row}"].alignment = wrap
    style_range(f"B{row}:C{row}", border=thin)

scenarios = [
    ("UAT on Performance Management navigation and access control", [
        ("Log in as HR Administrator / SuperAdmin and open the side menu.", "User lands on dashboard and Performance Management menu is visible."),
        ("Click 'Performance Management'.", "Dropdown/sub-menu appears with Setup, Manage Appraisals, Supervisor Evaluation, HOD Evaluation, Reports (as permitted)."),
        ("Expand Setup under Performance Management.", "Appraisal Settings, Rating Scales, Review Periods, Behavioral Items are listed. Focus Areas appears only when HR-defined approach is active."),
        ("Log in as a regular Staff user (ESS) and open Self Service.", "My Performance menu item is visible for the staff user."),
        ("Log in as a Supervisor user and open Performance Management.", "Supervisor Evaluation menu is accessible for appraisals where the user is supervisor."),
    ]),
    ("UAT on Appraisal Settings (organization approach switch)", [
        ("As Admin/HR, go to Performance Management > Setup > Appraisal Settings.", "Appraisal Settings page opens showing current appraisal approach and policy notes."),
        ("Select 'HR-defined performance areas & metrics' and click Save Settings.", "Success message is shown. Setting is saved as HR-defined."),
        ("Refresh menu / reopen Setup.", "Focus Areas menu is visible under Setup."),
        ("Change approach to 'Staff-defined goals, objectives & metrics' and Save.", "Success message is shown. Setting is saved as Staff-defined."),
        ("Refresh menu / reopen Setup.", "Focus Areas menu is hidden/unavailable for HR catalog setup."),
        ("Attempt to open Focus Areas URL directly while Staff-defined is active.", "User is redirected/blocked with a message that HR focus area setup is disabled for staff-defined mode."),
        ("Switch back to HR-defined and Save (prepare for next scenarios).", "Approach returns to HR-defined and Focus Areas menu becomes available again."),
    ]),
    ("UAT on Rating Scales setup", [
        ("Navigate to Performance Management > Setup > Rating Scales.", "Rating Scales list page is displayed."),
        ("Click Add / Create Rating Scale.", "Create Rating Scale form is displayed."),
        ("Enter Points, Rating Label, Description, Definition, Score Range, set Active, then Save.", "New rating scale is saved and appears in the list."),
        ("Edit an existing rating scale and update the label, then Save.", "Changes are saved successfully."),
        ("Deactivate a rating scale and Save.", "Scale status updates to inactive and remains listed appropriately."),
    ]),
    ("UAT on Review Periods setup", [
        ("Navigate to Performance Management > Setup > Review Periods.", "Review Periods list is displayed."),
        ("Click Create Review Period.", "Create form is displayed."),
        ("Enter Period Name, Start Date, End Date, set Active, then Save.", "Review period is created and listed."),
        ("Edit the review period dates and Save.", "Updated dates are saved successfully."),
        ("Attempt to create a review period with End Date before Start Date.", "Validation error is shown; record is not saved."),
    ]),
    ("UAT on HR-defined Focus Areas and Goals (HR-defined approach)", [
        ("Confirm Appraisal Settings is set to HR-defined.", "Approach shows HR-defined performance areas & metrics."),
        ("Go to Setup > Focus Areas and click Add Focus Area.", "Focus Area form opens."),
        ("Create Focus Area e.g. 'Financial Accuracy' with Weight 40%, optional Department/Designation, Active = Yes.", "Focus area is saved successfully."),
        ("Create additional focus areas so total weights for the intended employee scope equal 100%.", "Focus areas are listed with correct weights."),
        ("Open Goals for a Focus Area (list/goals icon).", "Goals list for that focus area is displayed."),
        ("Add a Goal with Strategic Objective, Performance Metric, Performance Target, Key Initiatives, Itemized Weighting.", "Goal is saved under the focus area."),
        ("Add at least 2 goals under different focus areas.", "Multiple goals are available for appraisal creation."),
        ("Edit a goal metric/target and Save.", "Goal updates successfully."),
    ]),
    ("UAT on Behavioral Items setup", [
        ("Navigate to Setup > Behavioral Items.", "Behavioral Items list is displayed."),
        ("Create a behavioral item with name, weight, and Active status.", "Behavioral item is saved and listed."),
        ("Edit and deactivate a behavioral item.", "Status updates successfully."),
    ]),
    ("UAT on Creating Appraisals in HR-defined mode", [
        ("Ensure Appraisal Settings = HR-defined and focus areas/goals exist for a test employee’s dept/designation.", "HR catalog metrics are ready."),
        ("Go to Performance Management > Manage Appraisals > Create.", "Appraisal create form opens with Review Period, Employee, Supervisor fields."),
        ("Select Review Period, Employee, Supervisor and Save/Create.", "Appraisal is created in Draft status."),
        ("Open the created appraisal (Show/View).", "Appraisal shows pre-loaded HR goals/metrics and behavioral items for scoring."),
        ("Verify staff ESS cannot see a 'Set Goals' action for this appraisal.", "Only Self Review / View actions are available (no staff goal setup)."),
        ("Attempt to create a duplicate appraisal for same employee + same review period (if validated).", "System prevents duplicate or warns appropriately."),
    ]),
    ("UAT on ESS Self Review (HR-defined appraisal)", [
        ("Log in as the appraised Staff user and go to Self Service > My Performance.", "My Performance appraisals list is displayed."),
        ("Confirm the draft/self-review appraisal is listed with correct period and supervisor.", "Appraisal row shows expected period, status, and supervisor."),
        ("Click 'Self Review' on the appraisal.", "Self Review form opens listing HR-defined focus areas, goals/metrics, and behavioral items."),
        ("Enter self ratings/weightings and comments for goals and behavioral items, then Save.", "Self review is saved; status moves to/remains Self Review; success message shown."),
        ("Re-open Self Review and confirm previously entered values are retained.", "Saved self ratings and comments are displayed."),
        ("Click Submit Self Review / Submit for supervisor.", "Success message indicates submission; supervisor can now review."),
        ("After submission (or once status moves beyond self_review), try editing self review again.", "Self review is locked / no longer editable."),
    ]),
    ("UAT on Supervisor Evaluation workflow", [
        ("Log in as the assigned Supervisor and open Performance Management > Supervisor Evaluation.", "List of appraisals awaiting supervisor review is displayed."),
        ("Confirm the submitted staff appraisal appears in the supervisor queue.", "Appraisal for the staff member is listed with Self Review / Supervisor Review status."),
        ("Open Supervisor Review for the appraisal.", "Supervisor review form shows staff self ratings and fields for supervisor ratings/comments."),
        ("Enter supervisor ratings/review weightings and comments for goals and behavioral items, then Save.", "Supervisor review is saved; status progresses to Supervisor Review."),
        ("Verify staff self ratings remain visible/read-only to supervisor.", "Self ratings are shown for reference and are not overwritten incorrectly."),
        ("Attempt supervisor review while logged in as a different non-supervisor user.", "Access is denied or appraisal is not listed."),
    ]),
    ("UAT on HOD Review, Finalize and Sign-off", [
        ("Log in as HOD/HR and open Performance Management > HOD Evaluation.", "Appraisals awaiting HOD review are listed."),
        ("Open HOD Review for the appraisal.", "HOD review form opens (comments / development / learning plans as applicable)."),
        ("Enter HOD comments and any development/learning plan entries, then Save.", "HOD review is saved; status moves to HOD Review."),
        ("Click Finalize on the appraisal.", "Appraisal status becomes Finalized; totals are recorded."),
        ("Perform Employee / Supervisor / HOD sign-off actions where available.", "Sign-off flags/dates are recorded successfully."),
        ("As staff, open My Performance > View on the finalized appraisal.", "Read-only appraisal details and scores are visible."),
    ]),
    ("UAT on Staff-defined appraisal approach (settings + appraisal shell)", [
        ("As Admin/HR, set Appraisal Settings to Staff-defined goals/objectives/metrics and Save.", "Staff-defined approach is active; success message shown."),
        ("Confirm Focus Areas HR catalog menu is hidden.", "Focus Areas is not available under Setup."),
        ("Create a new Appraisal for a staff member (Review Period, Employee, Supervisor).", "Appraisal is created in Draft without pre-loaded HR goal score rows. Message indicates staff will set their own goals."),
        ("Open the appraisal as Admin and verify goals/metrics section is empty (or awaiting staff setup), while behavioral items may still exist.", "No HR-defined goal scores are preloaded; appraisal shell is ready for staff goal setup."),
    ]),
    ("UAT on Staff Set Goals / Objectives / Metrics (ESS)", [
        ("Log in as the Staff user and go to My Performance.", "Appraisal list shows the staff-defined appraisal with a 'Set Goals' action."),
        ("Click 'Set Goals'.", "Set My Performance Goals form opens for the appraisal period."),
        ("Try saving with focus area weights totaling less than 100%.", "Validation error: focus area weights must total 100%."),
        ("Add Focus Area 1 (e.g. Delivery Excellence, Weight 60%) with at least one metric/criteria (objective, metric, target, weighting).", "Focus area and metric fields accept input."),
        ("Add Focus Area 2 (e.g. Stakeholder Engagement, Weight 40%) with at least one metric/criteria.", "Second focus area and metrics are added."),
        ("Click Save Goals & Continue to Self Rating.", "Goals are saved successfully and user is redirected to Self Review form."),
        ("Re-open Set Goals and confirm saved goals/metrics are loaded for editing.", "Previously saved staff goals are displayed."),
        ("Edit a metric target and Save again.", "Updated goals sync to the appraisal score rows successfully."),
    ]),
    ("UAT on Staff Self Review after setting own goals", [
        ("From My Performance, click Self Review on the staff-defined appraisal.", "Self Review form lists the staff-created focus areas/goals/metrics (and behavioral items if configured)."),
        ("Confirm there is an 'Edit My Goals' link while status is draft/self_review.", "Edit My Goals link is visible and opens the goals form."),
        ("Enter self ratings and comments against each staff-defined metric, then Save.", "Self review saves successfully."),
        ("Submit self evaluation for supervisor review.", "Submission succeeds; supervisor queue can pick up the appraisal."),
        ("After supervisor review has started / status advanced, attempt to edit goals again.", "Goal editing is blocked with an appropriate message."),
    ]),
    ("UAT on Supervisor + HR flow for staff-defined appraisals", [
        ("As Supervisor, open Supervisor Evaluation and locate the staff-defined appraisal.", "Appraisal appears after staff submission."),
        ("Open review and enter supervisor ratings against the staff-defined metrics, then Save.", "Supervisor ratings save; workflow progresses."),
        ("As HOD/HR, complete HOD review and Finalize.", "Appraisal is finalized successfully."),
        ("As HR, open Performance Reports and locate this employee/period.", "Finalized staff-defined appraisal is included in reports."),
    ]),
    ("UAT on Performance Reports", [
        ("As HR/Admin, open Performance Management > Reports > Department Report.", "Department report filter/page loads."),
        ("Select filters (department/period as applicable) and generate/view report.", "Department performance results for finalized appraisals are displayed."),
        ("Open Employee Report, select employee and generate.", "Employee appraisal report shows scores/totals."),
        ("Open Summary Report.", "Summary analytics/report for performance appraisals is displayed."),
        ("Attempt report access as a staff user without report permission.", "Access denied / menu not visible."),
    ]),
    ("UAT on Bulk appraisal upload (if used)", [
        ("As HR, go to Manage Appraisals and download the appraisal upload template.", "Template file downloads successfully."),
        ("With HR-defined approach active, upload a valid filled template for employees with matching catalog metrics.", "Appraisals are created and HR goals/behavioral items are preloaded."),
        ("Switch to Staff-defined approach, upload a valid template for other employees.", "Appraisals are created as staff-defined shells without HR goal preload."),
        ("Upload a template with invalid employee IDs.", "Rows fail with clear validation errors; valid rows may still process."),
    ]),
    ("UAT on Permissions, security and regression checks", [
        ("As Staff, attempt to open Admin Manage Appraisals / Appraisal Settings URLs.", "Access denied (403) or redirected; staff cannot change org approach."),
        ("As Staff on an HR-defined appraisal, confirm Set Goals is not offered.", "Only Self Review/View are available."),
        ("As Supervisor, confirm only assigned appraisals appear in Supervisor Evaluation.", "No unrelated employees’ appraisals are listed."),
        ("Create one HR-defined appraisal, then switch org setting to Staff-defined, and reopen the old appraisal.", "Existing appraisal retains HR-defined goals behavior (approach snapshot preserved)."),
        ("Create one Staff-defined appraisal, then switch org setting to HR-defined, and reopen the staff appraisal.", "Existing staff-defined appraisal still allows/shows staff goals path based on appraisal snapshot, not the new org default."),
        ("Verify finalized appraisals remain read-only across ESS and admin screens.", "No edits to scores/goals are allowed after finalize/close."),
    ]),
]

actual_options = '"Pending,As expected,Permission error,404- Not found,500- System error,505-Server error,Php error,Results shown differ with expected,403- Forbidden,408- Request time out"'
pass_options = '"Pass,Fail,Not executed,Suspended"'
dv_actual = DataValidation(type="list", formula1=actual_options, allow_blank=True)
dv_pass = DataValidation(type="list", formula1=pass_options, allow_blank=True)
ws.add_data_validation(dv_actual)
ws.add_data_validation(dv_pass)

current_row = 17
step_rows = []

for scenario_title, steps in scenarios:
    ws[f"A{current_row}"] = "Test Scenario"
    ws[f"A{current_row}"].font = font_bold
    ws.merge_cells(f"B{current_row}:L{current_row}")
    ws[f"B{current_row}"] = scenario_title
    ws[f"B{current_row}"].font = font_bold
    style_range(f"A{current_row}:L{current_row}", fill_scenario, thin, wrap)
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    headers = [
        ("A", "A", "Step #"),
        ("B", "C", "Step Details"),
        ("D", "E", "Expected Results"),
        ("F", "H", "Actual Results"),
        ("I", "J", "Pass / Fail / Not executed / Suspended"),
        ("K", "L", "Displayed Error/Success message"),
    ]
    for start, end, title in headers:
        if start != end:
            ws.merge_cells(f"{start}{current_row}:{end}{current_row}")
        ws[f"{start}{current_row}"] = title
        ws[f"{start}{current_row}"].font = font_bold
    style_range(f"A{current_row}:L{current_row}", fill_header, thin, center)
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    for idx, (detail, expected) in enumerate(steps, start=1):
        r = current_row
        step_rows.append(r)

        ws[f"A{r}"] = idx
        ws[f"A{r}"].font = font_normal
        ws[f"A{r}"].alignment = center
        ws[f"A{r}"].border = thin

        ws.merge_cells(f"B{r}:C{r}")
        ws[f"B{r}"] = detail
        ws[f"B{r}"].font = font_normal
        style_range(f"B{r}:C{r}", border=thin, alignment=wrap)

        ws.merge_cells(f"D{r}:E{r}")
        ws[f"D{r}"] = expected
        ws[f"D{r}"].font = font_normal
        style_range(f"D{r}:E{r}", border=thin, alignment=wrap)

        ws.merge_cells(f"F{r}:H{r}")
        ws[f"F{r}"] = "Pending"
        style_range(f"F{r}:H{r}", border=thin, alignment=center)

        ws.merge_cells(f"I{r}:J{r}")
        ws[f"I{r}"] = "Not executed"
        style_range(f"I{r}:J{r}", border=thin, alignment=center)

        ws.merge_cells(f"K{r}:L{r}")
        ws[f"K{r}"] = ""
        style_range(f"K{r}:L{r}", border=thin, alignment=wrap)

        dv_actual.add(f"F{r}")
        dv_pass.add(f"I{r}")
        ws.row_dimensions[r].height = 48
        current_row += 1

    current_row += 1

if step_rows:
    first = min(step_rows)
    last = max(step_rows)
    ws.conditional_formatting.add(
        f"F{first}:H{last}",
        FormulaRule(formula=[f'$F{first}="As expected"'], fill=fill_green),
    )
    ws.conditional_formatting.add(
        f"F{first}:H{last}",
        FormulaRule(
            formula=[
                f'OR($F{first}="Permission error",$F{first}="404- Not found",$F{first}="500- System error",'
                f'$F{first}="505-Server error",$F{first}="Php error",$F{first}="403- Forbidden",'
                f'$F{first}="408- Request time out",$F{first}="Results shown differ with expected")'
            ],
            fill=fill_orange,
        ),
    )
    ws.conditional_formatting.add(
        f"I{first}:J{last}",
        FormulaRule(formula=[f'$I{first}="Pass"'], fill=fill_green),
    )
    ws.conditional_formatting.add(
        f"I{first}:J{last}",
        FormulaRule(formula=[f'$I{first}="Fail"'], fill=fill_red),
    )

ws.freeze_panes = "A17"

# Cover
cover = wb.create_sheet("Cover", 0)
cover["A1"] = "MASTER UAT TEST DOCUMENTATION"
cover["A1"].font = Font(bold=True, size=18)
cover.merge_cells("A1:F1")
cover["A3"] = "Module"
cover["B3"] = "Performance Management"
cover["A4"] = "System"
cover["B4"] = "StawiHR"
cover["A5"] = "Document Type"
cover["B5"] = "User Acceptance Testing (UAT) Test Cases"
cover["A6"] = "Scope"
cover.merge_cells("B6:F8")
cover["B6"] = (
    "Full Performance Management module including: Appraisal Settings (HR-defined vs Staff-defined), "
    "Rating Scales, Review Periods, Focus Areas, Goals/KPIs, Behavioral Items, Appraisal creation "
    "(single & bulk), ESS My Performance, Staff goal setup, Self Review, Supervisor Evaluation, "
    "HOD Review, Finalize/Sign-off, and Performance Reports."
)
cover["B6"].alignment = wrap
cover["A10"] = "How to use"
cover["A10"].font = font_bold
cover.merge_cells("B10:F12")
cover["B10"] = (
    "1) Fill Tester Name and Date Tested on the Performance sheet.\n"
    "2) Execute each step in order within a Test Scenario.\n"
    "3) Record Actual Results using the dropdown.\n"
    "4) Mark Pass / Fail / Not executed / Suspended.\n"
    "5) Capture any displayed error/success message or notes in the last column."
)
cover["B10"].alignment = wrap
cover["A14"] = "Total Test Scenarios"
cover["B14"] = len(scenarios)
cover["A15"] = "Total Test Steps"
cover["B15"] = len(step_rows)
cover.column_dimensions["A"].width = 22
cover.column_dimensions["B"].width = 40
for col in ["C", "D", "E", "F"]:
    cover.column_dimensions[col].width = 18

# Legend
legend = wb.create_sheet("Legend")
legend["A1"] = "Actual Results dropdown values"
legend["A1"].font = font_bold
for i, v in enumerate([
    "Pending", "As expected", "Permission error", "404- Not found", "500- System error",
    "505-Server error", "Php error", "Results shown differ with expected", "403- Forbidden", "408- Request time out",
], start=2):
    legend[f"A{i}"] = v
legend["C1"] = "Pass/Fail dropdown values"
legend["C1"].font = font_bold
for i, v in enumerate(["Pass", "Fail", "Not executed", "Suspended"], start=2):
    legend[f"C{i}"] = v
legend["A14"] = "Suggested accounts for testing"
legend["A14"].font = font_bold
legend["A15"] = "Admin/HR"
legend["B15"] = "Configure settings, HR catalog, create appraisals, HOD/finalize, reports"
legend["A16"] = "Supervisor"
legend["B16"] = "Supervisor Evaluation queue and ratings"
legend["A17"] = "Staff"
legend["B17"] = "ESS My Performance, Set Goals (staff-defined), Self Review"
legend["A19"] = "Colour cues"
legend["A19"].font = font_bold
legend["A20"] = "As expected / Pass"
legend["B20"] = "Green"
legend["B20"].fill = fill_green
legend["A21"] = "Errors / mismatches"
legend["B21"] = "Orange"
legend["B21"].fill = fill_orange
legend["A22"] = "Fail"
legend["B22"] = "Red"
legend["B22"].fill = fill_red
legend.column_dimensions["A"].width = 28
legend.column_dimensions["B"].width = 70
legend.column_dimensions["C"].width = 28

out_path = r"D:\Development\School_layouts\stawihr_global2026\docs\Performance_Management_UAT_Test_Cases.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Scenarios: {len(scenarios)}")
print(f"Total steps: {len(step_rows)}")
